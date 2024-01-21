import streamlit as st
from PIL import Image
from datetime import datetime, timedelta
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
import shutil
import openpyxl
from openpyxl.styles import Alignment

def resize_image(input_path, output_path, desired_width_cm, desired_height_cm, transparent_bg=True):
    desired_width_px = int(desired_width_cm * 37.7952755906)
    desired_height_px = int(desired_height_cm * 37.7952755906)

    with Image.open(input_path) as img:
        resized_img = img.resize((desired_width_px, desired_height_px), Image.BICUBIC)

        if transparent_bg:
            new_img = Image.new("RGBA", (desired_width_px, desired_height_px), (0, 0, 0, 0))
            new_img.paste(resized_img, (0, 0), resized_img)
        else:
            new_img = Image.new("RGB", (desired_width_px, desired_height_px), "white")
            new_img.paste(resized_img, (0, 0), resized_img)

        new_img.save(output_path)

def resize_logo(input_path, output_path, desired_width, desired_height, transparent_bg=True):
    with Image.open(input_path) as logo_img:
        resized_logo = logo_img.resize((desired_width, desired_height), Image.BICUBIC)

        if transparent_bg:
            new_logo = Image.new("RGBA", (desired_width, desired_height), (0, 0, 0, 0))
            new_logo.paste(resized_logo, (0, 0), resized_logo)
        else:
            new_logo = Image.new("RGB", (desired_width, desired_height), "white")
            new_logo.paste(resized_logo, (0, 0), resized_logo)

        new_logo.save(output_path)

def generate_worksheet(sheet_name, start_row, start_column, num_checkmarks, output_path, output_firma1, output_firma2, output_logo, current_date, end_date, include_weekends):
    workbook = openpyxl.load_workbook(output_path)
    sheet = workbook[sheet_name]
    
    start_row += 1  # Adjusting to account for header row
    
    # Set the alignment style
    alignment = Alignment(horizontal="center", vertical="center")

    # Start row and column
    current_row = start_row
    current_column = start_column
    
    # Loop through dates and write data
    while current_date <= end_date:
        # Check if it's a weekday or weekends are included
        if current_date.weekday() < 5 or include_weekends:
            formatted_date = current_date.strftime('%Y-%m-%d')
            cell = sheet.cell(row=current_row, column=current_column, value=formatted_date)
            cell.alignment = alignment

            for _ in range(num_checkmarks):
                current_column += 1
                cell = sheet.cell(row=current_row, column=current_column, value="✔")
                cell.alignment = alignment

            # Add resized_firma1 to the next two cells (merged)
            resized_firma1_cell = sheet.cell(row=current_row, column=current_column + 1)
            sheet.merge_cells(start_row=current_row, start_column=current_column + 1, end_row=current_row, end_column=current_column + 2)
            sheet.add_image(OpenpyxlImage(output_firma1), f'{resized_firma1_cell.column_letter}{current_row}')

            # Add resized_firma2 to the next two cells (merged)
            resized_firma2_cell = sheet.cell(row=current_row, column=current_column + 3)
            sheet.merge_cells(start_row=current_row, start_column=current_column + 3, end_row=current_row, end_column=current_column + 4)
            sheet.add_image(OpenpyxlImage(output_firma2), f'{resized_firma2_cell.column_letter}{current_row}')

            current_row += 1
            current_column = start_column
        
        current_date += timedelta(days=1)

    # Add resized_logo to cell A1
    logo_cell = sheet.cell(row=1, column=1)
    sheet.add_image(OpenpyxlImage(output_logo), f'{logo_cell.column_letter}{logo_cell.row}')

    workbook.save(output_path)

# Streamlit app
st.title("Generador de formato")

# Step 1: Load the formato01.xlsx
file_formato01 = st.file_uploader("SUBA EL ARCHIVO formato01.xlsx", type=["xlsx"])
if file_formato01:
    # Step 2: Upload the images for firma1 and firma2
    st.write('Asegurese de que las firmas sean alargadas y de fondo blanco')
    firma1_image = st.file_uploader("SUBA LA FIRMA 1", type=["png", "jpg", "jpeg"])
    firma2_image = st.file_uploader("SUBA LA FIRMA 2", type=["png", "jpg", "jpeg"])
    
    # Step 3: Upload the logo image
    logo_image = st.file_uploader("SUBA EL LOGO DE SU EMPRESA", type=["png", "jpg", "jpeg"])

    # Step 4: Choose start_date, end_date, and include_weekends
    current_date = st.date_input("Elija el PRIMER dia del Mes que requiere", datetime.now())
    end_date = st.date_input("Elija la fecha final del formato", datetime.now() + timedelta(days=30))
    include_weekends = st.checkbox("DESEA INCLUIR FINES DE SEMANA", value=False)

    # Step 5: Write the path for the output directory
    output_directory = st.text_input("Enter the path for the output directory")

    # Step 6: Create the output excel button
    if st.button("Generate Excel"):
        if firma1_image and firma2_image and logo_image and output_directory:
            try:
                # Desired width and height for resizing images
                desired_width_cm = 4.5
                desired_height_cm = 0.72

                # Desired width and height for resizing the logo
                logo_width = 195
                logo_height = 93

                # Output paths for images and Excel file
                output_path = output_directory
                output_file = os.path.join(output_path, 'formatted_data2.xlsx')

                # File paths for images
                firma1 = os.path.join(output_path, 'firma1.png')
                firma2 = os.path.join(output_path, 'firma2.png')
                logo = os.path.join(output_path, 'logo.png')

                # Output paths for resized images
                output_firma1 = os.path.join(output_path, 'resized_firma1.png')
                output_firma2 = os.path.join(output_path, 'resized_firma2.png')
                output_logo = os.path.join(output_path, 'resized_logo.png')

                # Ensure the output directory exists
                os.makedirs(output_path, exist_ok=True)

                # Save uploaded images to the working folder
                Image.open(firma1_image).save(firma1)
                Image.open(firma2_image).save(firma2)
                Image.open(logo_image).save(logo)

                # Resize images and save them with a white background
                resize_image(firma1, output_firma1, desired_width_cm, desired_height_cm)
                resize_image(firma2, output_firma2, desired_width_cm, desired_height_cm)

                # Resize the logo image and save it with a transparent background
                resize_logo(logo, output_logo, logo_width, logo_height, transparent_bg=True)

                # Create a copy of the original file
                shutil.copy(file_formato01.name, output_file)

                # Specifications for each sheet
                sheet_parameters = [
                    {"name": "AREAS DE TRABAJO", "start_row": 8, "start_column": 2, "num_checkmarks": 10},
                    {"name": "AREA DE LA PLANTA", "start_row": 7, "start_column": 2, "num_checkmarks": 10},
                    {"name": "CUARTO FRIO", "start_row": 7, "start_column": 2, "num_checkmarks": 4},
                    {"name": "BAÑOS", "start_row": 6, "start_column": 2, "num_checkmarks": 6},
                    {"name": "TANQUES", "start_row": 7, "start_column": 2, "num_checkmarks": 8},
                    {"name": "OFICINA", "start_row": 7, "start_column": 2, "num_checkmarks": 10}
                ]

                for params in sheet_parameters:
                    generate_worksheet(params["name"], params["start_row"], params["start_column"], params["num_checkmarks"], output_file, output_firma1, output_firma2, output_logo, current_date, end_date, include_weekends)

                st.success("Excel file generated successfully.")
                st.text(f"Generated Excel file is stored at: {output_file}")

            except Exception as e:
                st.error(f"An error occurred: {e}")
